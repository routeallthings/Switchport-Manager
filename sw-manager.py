#!/usr/bin/env python

'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL netmiko (pip install netmiko)
INSTALL textfsm (pip install textfsm)
INSTALL openpyxl (pip install openpyxl)
INSTALL fileinput (pip install fileinput)
INSTALL xlhelper (python -m pip install git+git://github.com/routeallthings/xlhelper.git)
'''

#Module Imports (Native)
import re
import getpass
import os
import unicodedata
import csv
import threading
import time
import sys
from datetime import datetime
from datetime import date

#Module Imports (Non-Native)
try:
	import requests
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('Requests module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in requestsinstallstatus.upper() or "YES" in requestsinstallstatus.upper():
		os.system('python -m pip install requests')
		import requests
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of requests. Please install manually and retry"
		sys.exit()
try:
	import textfsm
except ImportError:
	textfsminstallstatus = fullpath = raw_input ('textfsm module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in textfsminstallstatus.upper() or "YES" in textfsminstallstatus.upper():
		os.system('python -m pip install textfsm')
		import textfsm
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of textfsm. Please install manually and retry"
		sys.exit()
try:
	import netmiko
	from netmiko import ConnectHandler
except ImportError:
	netmikoinstallstatus = fullpath = raw_input ('Netmiko module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in netmikoinstallstatus.upper() or "YES" in netmikoinstallstatus.upper():
		os.system('python -m pip install netmiko')
		import netmiko
		from netmiko import ConnectHandler
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netmiko. Please install manually and retry"
		sys.exit()
try:
	from openpyxl import load_workbook
	from openpyxl import workbook
	from openpyxl import Workbook
	from openpyxl.styles import Font, NamedStyle
except ImportError:
	openpyxlinstallstatus = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in openpyxlinstallstatus.lower():
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
		from openpyxl import workbook
		from openpyxl import Workbook
		from openpyxl.styles import Font, NamedStyle
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of openpyxl. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		
# PyPiWin32
try:
	import win32com.client
except ImportError:
	win32cominstallstatus = raw_input ('PyPiWin32 module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in win32cominstallstatus.lower():
		os.system('python -m pip install pypiwin32')
		os.system('python -m pip install pywin32')
		print 'You need to restart the script after installing win32com'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of PyPiWin32. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper		
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()
#######################################
#Functions
def DOWNLOAD_FILE(url,saveas):
    # NOTE the stream=True parameter
    r = requests.get(url, stream=True)
    with open(saveas, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024): 
            if chunk: # filter out keep-alive new chunks
                f.write(chunk)
def RestartPort(device,vendormac):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		vendormac = vendormac[:4] + '.' + vendormac[4:]
		showvendormac = 'show mac address-table | include ' + vendormac
		fullmaclist = sshnet_connect.send_command(showvendormac)
		if fullmaclist == '':
			sys.exit()
		fullmaclist = fullmaclist.split('\n')
		# Get Port #
		portlist = []
		for mac in fullmaclist:
			mac = " ".join(mac.split())
			maclist = mac.split(' ')
			portlist.append(maclist[3])
		# Restart Port #
		sshcommandset = []
		for port in portlist:
			iport = 'interface ' + port
			ishut = 'shutdown'
			inoshut = 'no shutdown'
			sshcommandset.append(iport)
			sshcommandset.append(ishut)
			sshcommandset.append(inoshut)
		FullOutput = sshnet_connect.send_config_set(sshcommandset)
		OutputLogp = loglocation + '\\' + devicehostname + '_portrestart.txt'
		if os.path.exists(OutputLogp):
			OutputLog = open(OutputLogp,'a+')
		else:
			OutputLog = open(OutputLogp,'w+')
		OutputLog.write('#################################################################\n')
		OutputLog.write('Start of Log\n')
		OutputLog.write('Current Start Time: ' + str(datetime.now()) + '\n')
		OutputLog.write(FullOutput)
		OutputLog.write('\n')
		OutputLog.close()
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error while gathering data with ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of DB'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''

def SetVLAN(device,vendormac,vendorvlan,vendorvlanmod,vendortemplate):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		vendormac = vendormac[:4] + '.' + vendormac[4:]
		showvendormac = 'show mac address-table | include ' + vendormac
		fullmaclist = sshnet_connect.send_command(showvendormac)
		if fullmaclist == '':
			sys.exit()
		fullmaclist = fullmaclist.split('\n')
		# Get Port #
		vlanportlist = []
		for mac in fullmaclist:
			mac = " ".join(mac.split())
			maclist = mac.split(' ')
			vlanport = maclist[0] + ',' + maclist[3]
			vlanportlist.append(vlanport)
		# Build list of changes
		findtrunk = 'show interface trunk | i trunking'
		trunklist = sshnet_connect.send_command(findtrunk)
		sshcommandset = []
		portcompliance = []
		for vlanport in vlanportlist:
			vlanport = vlanport.split(',')
			vlannumber = vlanport[0]
			vlaninterface = vlanport[1]
			if vlaninterface in trunklist:
				notrunk = 0
			else:
				notrunk = 1
			if not vlannumber in vendorvlan:
				if modifyvlan == 0 and notrunk == 1:
					print 'Port is out of compliance: ' + devicehostname + ',' + vlaninterface + ',' + vlannumber
				if modifyvlan == 1 and notrunk == 1:
					iport = 'interface ' + vlaninterface
					ivlan = 'switchport access vlan ' + vendorvlanmod
					itemplate = 'source template ' + vendortemplate
					sshcommandset.append(iport)
					sshcommandset.append(ivlan)
					sshcommandset.append(itemplate)
		if modifyvlan == 1 and sshcommandset != []:
			FullOutput = sshnet_connect.send_config_set(sshcommandset)
			OutputLogp = loglocation + '\\' + devicehostname + '_portvlan.txt'
			if os.path.exists(OutputLogp):
				OutputLog = open(OutputLogp,'a+')
			else:
				OutputLog = open(OutputLogp,'w+')
			OutputLog.write('#################################################################\n')
			OutputLog.write('Start of Log\n')
			OutputLog.write('Current Start Time: ' + str(datetime.now()) + '\n')
			OutputLog.write(FullOutput)
			OutputLog.write('\n')
			OutputLog.close()
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''

def SetTemplate(device,vendorvlan,vendortemplate):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		showinterfacelist = 'sh ip int br | e Vlan|Tunnel|Loopback|Port-channel|GigabitEthernet0/0|Interface'
		fullinterfacelist = sshnet_connect.send_command(showinterfacelist)
		if fullinterfacelist == '':
			sys.exit()
		fullinterfacelist = fullinterfacelist.split('\n')
		for fullint in fullinterfacelist:
			sshcommandset = []
			sourcevlan = ''
			sourcetemplate = ''
			settemplatecmd = 'sh run interface ' + fullint + ' | i (source template)'
			sourcetemplatessh = sshnet_connect.send_command(settemplatecmd)
			sourcetemplate = re.search('\S+\s+\S+\s+(\S+)',sourcetemplatessh)
			sourcetemplate = sourcetemplate.group(1)
			setvlancmd = 'sh der interface ' + fullint + ' | i (access vlan)'
			sourcevlanssh = sshnet_connect.send_command(settemplatecmd)
			sourcevlan = re.search('switchport\s+access\s+vlan\s+(\S+)',sourcevlanssh)
			sourcevlan = sourcevlan.group(1)
			if sourcetemplate != vendortemplate and sourcevlan == vendorvlan:
				iport = 'interface ' + vlaninterface
				itemplate = 'source template ' + vendortemplate
				sshcommandset.append(iport)
				sshcommandset.append(itemplate)
				FullOutput = sshnet_connect.send_config_set(sshcommandset)
				OutputLogp = loglocation + '\\' + devicehostname + '_portvlan.txt'
				if os.path.exists(OutputLogp):
					OutputLog = open(OutputLogp,'a+')
				else:
					OutputLog = open(OutputLogp,'w+')
				OutputLog.write('#################################################################\n')
				OutputLog.write('Start of Log\n')
				OutputLog.write('Current Start Time: ' + str(datetime.now()) + '\n')
				OutputLog.write(FullOutput)
				OutputLog.write('\n')
				OutputLog.close()
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''

def MacIPCompare(device,vendorvlan,modifyipport,modifyvendorvlan):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		# Get Macs
		showmaclist = 'show mac address vlan ' + str(vendorvlan) + ' | e Vl|CPU|--|Address|Po'
		fullmaclist = sshnet_connect.send_command(showmaclist)
		if fullmaclist == '':
			sys.exit()
		fullmaclist = fullmaclist.split('\n')
		maclist = []
		for fullmac in fullmaclist:
			if fullmac == '':
				continue
			macentry = {}
			macreg = re.search('\S+\s+(\S+)\s+\S+\s+(\S+).*',fullmac)
			indmac = macreg.group(1)
			indint = macreg.group(2)
			macentry['mac'] = indmac
			macentry['int'] = indint
			maclist.append(macentry)
		# Get VRFs
		vrflist = []
		showvrfcmd = 'show vrf | i ipv4'
		showvrf = sshnet_connect.send_command(showvrfcmd)
		showvrf = showvrf.split('\n')
		try:
			vrfcount = 0
			for v in showvrf:
				if not 'Mgmt-vrf' in v:
					vrf1 = re.search('(\S+)\s+.*',v)
					vrf1 = vrf1.group(1)
					vrflist.append(vrf1)
					vrfcount = vrfcount + 1
			if vrfcount == 0:
				vrffound = 0
			else:
				vrffound = 1
		except:
			vrffound = 0
		# Get IPs in VLAN
		arpfoundips = 0
		if vrffound == 1:
			for vrf in vrflist:
				iparpcmd = 'show ip arp vrf ' + vrf + ' vlan ' + vendorvlan + ' | e Address'
				iparplist = sshnet_connect.send_command(iparpcmd)
				if iparplist == '':
					continue
				else:
					arpfoundips = 1
					break
		if not arpfoundips == 1:
			iparpcmd = 'show ip arp vlan ' + vendorvlan + ' | e Address'
			iparplist = sshnet_connect.send_command(iparpcmd)
		iparplist = iparplist.split('\n')
		arplist = []
		for iparp in iparplist:
			if iparp == '':
				continue
			arpentry = {}
			iparpreg = re.search('\S+\s+(\S+)\s+\S+\s+(\S+).*',iparp)
			iparpip = iparpreg.group(1)
			iparpmac = iparpreg.group(2)
			arpentry['ip'] = iparpip
			arpentry['mac'] = iparpmac
			arplist.append(arpentry)
		# Compare the mac list and the arp table
		sshcommandset = []
		for mac in maclist:
			foundarpmac = 0
			indmac = mac.get('mac')
			indint = mac.get('int')
			for arp in arplist:
				arpmac = arp.get('mac')
				arpip = arp.get('ip')
				if arpmac == indmac:
					foundarpmac = 1
					break
			if foundarpmac == 0 and modifyipport == 0:
				print devicehostname + ' - No IP found for the mac address ' + indmac + ' on the port ' + indint
			if foundarpmac == 0 and modifyipport == 1:
				iport = 'interface ' + indint
				ivlan = 'switchport access vlan ' + modifyvendorvlan
				sshcommandset.append(iport)
				sshcommandset.append(ivlan)
		if modifyipport == 1 and sshcommandset != []:
			FullOutput = sshnet_connect.send_config_set(sshcommandset)
			OutputLogp = loglocation + '\\' + devicehostname + '_noip.txt'
			if os.path.exists(OutputLogp):
				OutputLog = open(OutputLogp,'a+')
			else:
				OutputLog = open(OutputLogp,'w+')
			OutputLog.write('#################################################################\n')
			OutputLog.write('Start of Log\n')
			OutputLog.write('Current Start Time: ' + str(datetime.now()) + '\n')
			OutputLog.write(FullOutput)
			OutputLog.write('\n')
			OutputLog.close()
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
			
def ExportVLANs(device):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		devicehostnames.append(devicehostname)
		# Get VRFs
		vrflist = []
		showvrfcmd = 'show vrf | i ipv4'
		showvrf = sshnet_connect.send_command(showvrfcmd)
		showvrf = showvrf.split('\n')
		try:
			vrfcount = 0
			for v in showvrf:
				if not 'Mgmt-vrf' in v:
					vrf1 = re.search('(\S+)\s+.*',v)
					vrf1 = vrf1.group(1)
					vrflist.append(vrf1)
					vrfcount = vrfcount + 1
			if vrfcount == 0:
				vrffound = 0
			else:
				vrffound = 1
		except:
			vrffound = 0
		# Get Interface Dictionary
		#intbrlist = []
		#showintbrcmd = 'show ip int br | e unassigned|Method'
		#showintbrssh = sshnet_connect.send_command(showintbrcmd)
		#showintbrssh = showintbrssh.split('\n')
		#for intbr in showintbrssh:
		#	intbrreg = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)',intbr)
		#	intbrint = intbrreg.group(1)
		#	intbrvrfcmd = 'show running int ' + intbrint + ' | i forwarding'
		#	if not intbrvrfcmd = ''
		#		intbrvrf = 
		
		# Get Interfaces List
		showinterfacelistcmd = 'sh ip int br | e Vlan|Tunnel|Loopback|Port-channel|GigabitEthernet0/0|Interface'
		fullinterfacelistbr = sshnet_connect.send_command(showinterfacelistcmd)
		if fullinterfacelistbr == '':
			sys.exit()
		fullinterfacelistbr = fullinterfacelistbr.split('\n')
		fullinterfacelist = []
		for fullinterfaceline in fullinterfacelistbr:
			if fullinterfaceline == '':
				continue
			fullinterfacedict = {}
			fullinterfacereg = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)',fullinterfaceline)
			fullinterface = fullinterfacereg.group(1)
			if 'Te' in fullinterface or 'Fo' in fullinterface:
				if 'Te' in fullinterface:
					fullintspeed = 'TenGigabitEthernet'
				if 'Fo' in fullinterface:
					fullintspeed = 'FortyGigabitEthernet'
				fullintname = re.search('[A-Za-z]+(\S+)',fullinterface)
				fullintname = fullintspeed + fullintname.group(1)
			else:
				fullintname = fullinterface
			fullinterfacedict['int'] = fullintname
			fullinterfacedict['status'] = fullinterfacereg.group(5)
			fullinterfacelist.append(fullinterfacedict)
		# Build MAC DB
		macaddlist = []
		showmaccmd = 'show mac address-table | i /'
		showmaclist = sshnet_connect.send_command(showmaccmd)
		showmaclist = showmaclist.split('\n')
		for mac in showmaclist:
			if not mac == '':
				try:
					vlanmac = ''
					addmac = ''
					intmac = ''
					macadddict = {}
					fullmacline = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+)',mac)
					vlanmac = fullmacline.group(1)
					addmac = fullmacline.group(2)
					intmac = fullmacline.group(4)
					macadddict['vlan'] = vlanmac
					macadddict['add'] = addmac
					macadddict['int'] = intmac
					macaddlist.append(macadddict)
				except:
					continue
		# Build POE DB
		fullpoeinterfacecmd = 'show power inline | i /'
		fullpoeinterfacelist = sshnet_connect.send_command(fullpoeinterfacecmd)
		fullpoeinterfacelist = fullpoeinterfacelist.split('\n')
		fullpoeinterfacedictlist = []
		for poeint in fullpoeinterfacelist:
			if poeint == '':
				continue
			fullpoeinterfacedict = {}
			fullpoeinterfaceshow = ''
			try:
				fullpoeinterfaceshow = re.search('(\S+)\s+\S+\s+(\S+)\s+(\S+)\s+(\S+\s?\S+?\s?\S+?)\s+\S+\.*',poeint)
				intpoename =  fullpoeinterfaceshow.group(1)
				if 'on' in fullpoeinterfaceshow.group(2):
					intpoestatus = fullpoeinterfaceshow.group(2)
					intpoewatts = fullpoeinterfaceshow.group(3)
					intpoedevice = fullpoeinterfaceshow.group(4)
				else:
					intpoestatus = ''
					intpoewatts = ''
					intpoedevice = ''			
				fullpoeinterfacedict['int'] = intpoename
				fullpoeinterfacedict['status'] = intpoestatus
				fullpoeinterfacedict['watts'] = intpoewatts
				fullpoeinterfacedict['device'] = intpoedevice
				fullpoeinterfacedictlist.append(fullpoeinterfacedict)
			except:
				continue
		# Build CDP DB
		fullcdpinterfacecmd = 'show cdp nei det'
		fullcdpinterfacelist = sshnet_connect.send_command(fullcdpinterfacecmd)
		fullcdpinterfacelist = fullcdpinterfacelist.split('-------------------------')
		fullcdpinterfacedictlist = []
		for cdpintline in fullcdpinterfacelist:
			try:
				cdpdeviceid = ''
				cdpplatformid = ''
				cdpinterfaceid = ''
				cdpint = cdpintline.split('\n')
				fullcdpinterfacedict = {}
				try:
					for cdpin in cdpint:
						if 'Device ID' in cdpin:
							cdpdeviceid = re.search('.*Device ID: (.*)',cdpin)
							cdpdeviceid = cdpdeviceid.group(1)
						if 'Platform' in cdpin:
							cdpplatformid = re.search('.*Platform: (.*),.*',cdpin)
							cdpplatformid = cdpplatformid.group(1)
						if 'Interface' in cdpin:
							cdpinterfaceid = re.search('.*Interface: (\S+),.*',cdpin)
							cdpinterfaceid = cdpinterfaceid.group(1)
				except:
					continue
				fullcdpinterfacedict['int'] = cdpinterfaceid
				fullcdpinterfacedict['platform'] = cdpplatformid
				fullcdpinterfacedict['device'] = cdpdeviceid
				fullcdpinterfacedictlist.append(fullcdpinterfacedict)
			except:
				continue
		# Get VRF to VLAN assignment
		if vrffound == 1:
			vrfassignmentlist = []
			shvrfintcmd = 'show ip vrf interfaces | e Interface'
			shvrfint = sshnet_connect.send_command(shvrfintcmd)
			shvrfint = shvrfint.split('\n')
			for shvrf in shvrfint:
				if shvrf == '':
					continue
				try:
					shvrfdict = {}
					shvrfreg = re.search('[A-Za-z][A-Za-z](\d+)\s+\S+\s+(\S+)\s+.*',shvrf) 
					shvrfvlan = shvrfreg.group(1)
					shvrfvrf = shvrfreg.group(2)
					shvrfdict['vrf'] = shvrfvrf
					shvrfdict['vlan'] = shvrfvlan
					vrfassignmentlist.append(shvrfdict)
				except:
					continue
		# Build ARP table per VRF
		fullarptablelist = []
		if vrffound == 0:
			arpvrfcmd = 'show ip arp | e Protocol|Incomplete'
			arpvrflist = sshnet_connect.send_command(arpvrfcmd)
			arpvrflist = arpvrflist.split('\n')
			for arpentryline in arpvrflist:
				try:
					arpentrydict = {}
					arpentryip = ''
					arpentrymac = ''
					arpentryvlan = ''
					arpentry = re.search('\S+\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+(\S+).*',arpentryline)
					arpentryip = arpentry.group(1)
					arpentrymac = arpentry.group(2)
					arpentryvlan = arpentry.group(3)
					arpentrydict['ip'] = arpentryip
					arpentrydict['mac'] = arpentrymac
					arpentrydict['vlan'] = arpentryvlan
					arpentrydict['vrf'] = ''
					fullarptablelist.append(arpentrydict)
				except:
					continue
		if vrffound == 1:
			# Get Global Table
			arpvrfcmd = 'show ip arp | e Protocol|Incomplete'
			arpvrflist = sshnet_connect.send_command(arpvrfcmd)
			arpvrflist = arpvrflist.split('\n')
			for arpentryline in arpvrflist:
				try:
					arpentrydict = {}
					arpentryip = ''
					arpentrymac = ''
					arpentryvlan = ''
					arpentry = re.search('\S+\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+(\S+).*',arpentryline)
					arpentryip = arpentry.group(1)
					arpentrymac = arpentry.group(2)
					arpentryvlan = arpentry.group(3)
					arpentrydict['ip'] = arpentryip
					arpentrydict['mac'] = arpentrymac
					arpentrydict['vlan'] = arpentryvlan
					arpentrydict['vrf'] = ''
					fullarptablelist.append(arpentrydict)
				except:
					continue
			# Get VRF Tables
			for vrfin in vrflist:
				arpvrfcmd = 'show ip arp vrf ' + vrfin + ' | e Protocol|Incomplete'
				arpvrflist = sshnet_connect.send_command(arpvrfcmd)
				arpvrflist = arpvrflist.split('\n')
				for arpentryline in arpvrflist:
					try:
						arpentrydict = {}
						arpentryip = ''
						arpentrymac = ''
						arpentryvlan = ''
						arpentry = re.search('\S+\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+(\S+).*',arpentryline)
						arpentryip = arpentry.group(1)
						arpentrymac = arpentry.group(2)
						arpentryvlan = arpentry.group(3)
						arpentrydict['ip'] = arpentryip
						arpentrydict['mac'] = arpentrymac
						arpentrydict['vlan'] = arpentryvlan
						arpentrydict['vrf'] = vrfin
						fullarptablelist.append(arpentrydict)
					except:
						continue
		# Start of Interface Stat		
		for fullint in fullinterfacelist:
			# Get Basic Variables
			intname = fullint.get('int')
			intstatus = fullint.get('status')
			# Reset Variables
			interfacedictionary = {}
			intvlan = ''
			intmac = ''
			maccompany = ''
			intip = ''
			intvrf = ''
			inttemplate = ''
			cdphostname = ''
			cdpplatform = ''
			intpoestatus = ''
			intpoewatts = ''
			intpoedevice = ''
			# Start			
			#fullintreg = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)',fullint)
			#try:
			#	intname = fullintreg.group(1)
			#except:
			#	intname = ''
			#try:
			#	intstatus = fullintreg.group(5)
			#except:
			#	intstatus = ''
			# Create short interface name
			if re.match('[A-Za-z][A-Za-z]\d+/\d+/\d+',intname):
				shortintname = intname
			else:
				shortintname = re.search('([A-Za-z][A-Za-z])[A-Za-z]+(\S+)',intname)
				shortintname = shortintname.group(1) + shortintname.group(2)
			# Data gathering if port is up
			if 'up' in intstatus:
				showmac = ''
				intmac = ''
				intip = ''
				for macd in macaddlist:
					if macd['int'] == shortintname:
						intmac = macd['add']
						try:
							mac_company_mac = str(intmac[0:7].replace('.','')).upper()
							for line in maclookupdb:
								if line.startswith(mac_company_mac):
									linev = line.replace('\n','').replace('\t',' ')
									maccompany = re.search(r'^[A-Z0-9]{6}\s+\(base 16\)\s+(.*)',linev).group(1)
								if maccompany == '' or maccompany == None:
									maccompany = 'Unknown'
						except:
							maccompany = 'Unknown'
						break
				#
				#showmaccmd = 'show mac address-table interface ' + intname + ' | include /'
				#showmac = sshnet_connect.send_command(showmaccmd)
				#showmac = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+)',showmac)
				#try:
				#	intmac = showmac.group(2)
				#	if not intmac == '':
				#		try:
				#			mac_company_mac = str(intmac[0:7].replace('.','')).upper()
				#			for line in maclookupdb:
				#				if line.startswith(mac_company_mac):
				#					linev = line.replace('\n','').replace('\t',' ')
				#					maccompany = re.search(r'^[A-Z0-9]{6}\s+\(base 16\)\s+(.*)',linev).group(1)
				#				if maccompany == '' or maccompany == None:
				#					maccompany = 'Unknown'
				#		except:
				#			maccompany = 'Unknown'
				#	else:
				#		maccompany = ''
				#except:
				#	intmac = 'No MAC or Error'
				#	intip = ''
				#	maccompany = ''
				#	continue
				# If company is Cisco, do CDP CHECK
				#if 'Cisco' in maccompany:
				#	cdphostname = ''
				#	cdpplatform = ''
				#	cdpcheckcmd = 'show cdp neighbor ' + intname + ' detail | e cdp'
				#	showcdp = sshnet_connect.send_command(cdpcheckcmd)
				#	showcdp = showcdp.split('\n')
				#	for cdpr in showcdp:
				#		if 'Device ID:' in cdpr:
				#			cdprreg = re.search('.*Device ID: (\S+).*',cdpr)
				#			cdphostname = cdprreg.group(1)
				#		if 'Platform:' in cdpr:
				#			cdprreg = re.search('.*Platform: (.*),.*',cdpr)
				#			cdpplatform = cdprreg.group(1)
				#else:
				#	cdphostname = ''
				#	cdpplatform = ''
				
				
				# Add any CDP information
				for cdpinfo in fullcdpinterfacedictlist:
					if cdpinfo.get('int') == intname:
						cdphostname = cdpinfo.get('device')
						cdpplatform = cdpinfo.get('platform')
						break
				
				# VRF check
				#try:
				#	if vrffound == 0:
				#		showipcmd = 'show ip arp | i ' + intmac
				#		showip = sshnet_connect.send_command(showipcmd)
				#	if vrffound == 1:
				#		for v in vrflist:
				#			showipcmd = 'show ip arp vrf ' + v + ' | i ' + intmac
				#			showip = sshnet_connect.send_command(showipcmd)
				#			if showip != '':
				#				intvrf = v
				#				break
				#		if showip == '':
				#			showipcmd = 'show ip arp | i ' + intmac
				#			showip = sshnet_connect.send_command(showipcmd)
				#			intvrf = ''
				#	showip = re.search('\S+\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}).*',showip)
				#	intip = showip.group(1)
				#except:
				#	intip = ''
				
				# ARP Check
				for arpline in fullarptablelist:
					if arpline.get('mac') == intmac:
						intip = arpline.get('ip')
						intvrf = arpline.get('vrf')
			else:
				intmac = ''
				intip = ''
				intvrf = ''
				maccompany = ''
				cdphostname = ''
				cdpplatform = ''
			showrunintcmd = 'show running-config interface ' + intname
			showrunint = sshnet_connect.send_command(showrunintcmd)
			#if 'Invalid' in showrunint:
			#	showrunintcmd = 'show running-config interface ' + intname
			#	showrunint = sshnet_connect.send_command(showrunintcmd)
			try:
				showrunint = showrunint.split('\n')
				for int in showrunint:
					if 'switchport access' in int:
						intvlan = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+).*',int)
						try:
							intvlan = intvlan.group(4)
						except:
							intvlan = 'Error'
					if 'switchport mode trunk' in int:
						intvlan = 'Trunk'
					if 'source template' in int:
						inttemplate = re.search('(\S+)\s+(\S+)\s+(\S+).*',int)
						try:
							inttemplate = inttemplate.group(3)
						except:
							inttemplate = ''
				if intvlan == '':
					showdervlancmd = 'show derived-config interface ' + intname + ' | i (switchport access vlan)|(switchport mode trunk)'
					showdervlanres = sshnet_connect.send_command(showdervlancmd)
					showdervlanres = showdervlanres.split('\n')
					for showder in showdervlanres:
						if 'switchport access' in showder:
							intvlan = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+).*',showder)
							try:
								intvlan = intvlan.group(4)
							except:
								intvlan = 'Error'
						if 'switchport mode trunk' in showder:
							intvlan = 'Trunk'
			except:			
				intvlan = ''
				inttemplate = ''
			# Get VRF for any missing
			if intvrf == '' and vrffound == 1:
				for vrfdictline in vrfassignmentlist:
					if intvlan == vrfdictline.get('vlan'):
						intvrf = vrfdictline.get('vrf')
						break

			# Get POE status on the port
			#showpoeintcmd = 'show power inline ' + intname + ' | e --|Watts|Power'
			#try:
			#	showpoeint = sshnet_connect.send_command(showpoeintcmd)
			#	showpoeint = re.search('\S+\s+\S+\s+(\S+)\s+(\S+)\s+(\S+).*',showpoeint)
			#	if 'on' in showpoeint.group(1):
			#		intpoestatus = showpoeint.group(1)
			#		intpoewatts = showpoeint.group(2)
			#		intpoedevice = showpoeint.group(3)
			#	else:
			#		intpoestatus = ''
			#		intpoewatts = ''
			#		intpoedevice = ''				
			#except:
			#	intpoestatus = ''
			#	intpoewatts = ''
			#	intpoedevice = ''
			# Get POE status from dictionary
			for fullpoeint in fullpoeinterfacedictlist:
				if shortintname == fullpoeint.get('int'):
					intpoestatus = fullpoeint.get('status')
					intpoewatts = fullpoeint.get('watts')
					intpoedevice = fullpoeint.get('device')
			# Build Dictionary
			interfacedictionary['Hostname'] = devicehostname
			interfacedictionary['Interface'] = intname
			interfacedictionary['VLAN'] = intvlan
			interfacedictionary['Status'] = intstatus
			interfacedictionary['MacAddress'] = intmac
			interfacedictionary['MacCompany'] = maccompany
			interfacedictionary['IPAddress'] = intip
			interfacedictionary['VRF'] = intvrf
			interfacedictionary['Template'] = inttemplate
			interfacedictionary['CDPHostname'] = cdphostname
			interfacedictionary['CDPPlatform'] = cdpplatform
			interfacedictionary['POEStatus'] = intpoestatus
			interfacedictionary['POEWatts'] = intpoewatts
			interfacedictionary['POEDevice'] = intpoedevice
			finalinterfacelist.append(interfacedictionary) 
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''

def DOWNLOAD_FILE(url,saveas):
    # NOTE the stream=True parameter
    r = requests.get(url, stream=True)
    with open(saveas, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024): 
            if chunk: # filter out keep-alive new chunks
                f.write(chunk)
			
def HealthCheck(device):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	### FSM Templates ###
	# FSM Show Interface
	if "cisco_ios" in devicetype:
		fsmshowinturl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interfaces_health.template"
	if "cisco_xe" in devicetype:
		fsmshowinturl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interfaces_health.template"
	if "cisco_nxos" in devicetype:
		fsmshowinturl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_nxos_show_interfaces_health.template"
	fsmtemplatename = devicetype + '_fsmshowint_health.fsm'
	if not os.path.isfile(fsmtemplatename):
		DOWNLOAD_FILE(fsmshowinturl, fsmtemplatename)
	fsmtemplatenamefile = open(fsmtemplatename)
	fsminttemplate = textfsm.TextFSM(fsmtemplatenamefile)
	tempfilelist.append(fsmtemplatename)
	fsmtemplatenamefile.close()
	# FSM Show Temperature
	if "cisco_ios" in devicetype:
		fsmshowtempurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_temp_health.template"
	if "cisco_xe" in devicetype:
		fsmshowtempurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_temp_health.template"
	if "cisco_nxos" in devicetype:
		fsmshowtempurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_nxos_show_temp_health.template"	
	fsmtemplatename = devicetype + '_fsmshowtemp_health.fsm'
	if not os.path.isfile(fsmtemplatename):
		DOWNLOAD_FILE(fsmshowtempurl, fsmtemplatename)
	fsmtemplatenamefile = open(fsmtemplatename)
	fsmtemptemplate = textfsm.TextFSM(fsmtemplatenamefile)
	tempfilelist.append(fsmtemplatename)
	fsmtemplatenamefile.close()
	#Start Connection
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		devicehostnames.append(devicehostname)
		#Show Interfaces
		showinterface = 'show interface'
		sshcommand = showinterface
		sshresult = sshnet_connect.send_command(sshcommand)
		hcshowint = fsminttemplate.ParseText(sshresult)
		#Parse through each interface looking for issues
		healthcheckcsv = []
		for hcshowintsingle in hcshowint:
			hcinterfacename = hcshowintsingle[0].encode('utf-8')
			if not 'notconnect' in hcshowintsingle[2]:
				# Look for duplexing issues
				if 'Half-duplex' in hcshowintsingle[6]:
					hcerror = 'Duplex Mismatch'
					hcdescription = hcinterfacename + ' is showing as half-duplex. If this is by design please ignore.'
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
				if '10Mb/s' in hcshowintsingle[7]:
					hcerror = 'Duplex Mismatch'
					hcdescription = hcinterfacename + ' is showing as 10Mb/s. If this is by design please ignore.'
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
				# Look for interface counter errors
				# Input Errors
				hcshowintsingleint = hcshowintsingle[8]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)
				if hcshowintsingleint > 20:
					hcerror = 'Input Errors'
					hcinterfacecounter = hcshowintsingle[8]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' input errors. Usually indicative of a bad link (cabling and/or optic failure).'
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
				# CRC errors
				hcshowintsingleint = hcshowintsingle[9]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)			
				if hcshowintsingleint > 20:
					hcerror = 'CRC Errors'
					hcinterfacecounter = hcshowintsingle[9]
					hcinterfacecounter = hcinterfacecounter
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' CRC errors. Usually indicative of incorrect duplexing settings or a bad link (cabling and/or optic failure).'
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
				# Output errors
				hcshowintsingleint = hcshowintsingle[10]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)		
				if hcshowintsingleint > 10000:
					hcerror = 'Saturated Link'
					hcinterfacecounter = hcshowintsingle[10]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' output errors. This is usually indicative of a saturated interface.  '
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
				# Collisions
				hcshowintsingleint = hcshowintsingle[11]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)
				if hcshowintsingleint > 20:
					hcerror = 'Shared Medium'
					hcinterfacecounter = hcshowintsingle[11]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' collisions.  '
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))		
				# Interface resets
				hcshowintsingleint = hcshowintsingle[12]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)			
				if hcshowintsingleint > 20:
					hcerror = 'Interface Reset Count'
					hcinterfacecounter = hcshowintsingle[12]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' interface resets. '
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
		#Show Temperature
		try:
			if 'cisco_ios' in sshdevicetype.lower() or 'cisco_xe' in sshdevicetype.lower():
				sshcommand = showtemp
				sshresult = sshnet_connect.send_command(sshcommand)
				hcshowtemp = fsmtemptemplate.ParseText(sshresult)
				hctempdegrees = hcshowtemp[0]
				hctempdegrees = hctempdegrees[0]
				hctempdegrees = hctempdegrees.encode('utf-8')
				hctempdegreesint = int(hctempdegrees)
				if hctempdegreesint > 45:
					hcerror = 'Temperature Alert'
					hcdescription = 'Temperature has been recorded at ' + hctempdegrees + ' Celsius. Please lower the temperature for the surrounding environment '
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
			if 'cisco_nxos' in sshdevicetype.lower():
				sshcommand = showtemp_nxos
				sshresult = sshnet_connect.send_command(sshcommand)
				hcshowtemp = fsmtemptemplate.ParseText(sshresult)
				hctempdegrees = hcshowtemp[0]
				hctempdegrees = hctempdegrees[0]
				hctempdegrees = hctempdegrees.encode('utf-8')
				hctempdegreesint = int(hctempdegrees)
				if hctempdegreesint > 45:
					hcerror = 'Temperature Alert'
					hcdescription = 'Temperature has been recorded at ' + hctempdegrees + ' Celsius. Please lower the temperature for the surrounding environment '
					healthcheckcsv.append ((devicehostname + ',' + hcerror + ',' + hcdescription))
		except:
			pass
		# Exit SSH
		sshnet_connect.disconnect()
		# Parse list into dictionary/list
		saveresultslistsplit = []
		for saveresultsrow in healthcheckcsv:
			saveresultslistsplit.append(saveresultsrow.strip().split(','))
		saveresultslistsplit = [saveresultslistsplit[i:i+3] for i in range(0,len(saveresultslistsplit),3)]
		for saveresultsplitrow in saveresultslistsplit:
			for saveresultssplitrow2 in saveresultsplitrow:
				tempdict = {}
				tempdict['Hostname'] = saveresultssplitrow2[:1][0]
				tempdict['Error'] = saveresultssplitrow2[1:][0]
				tempdict['Description'] = saveresultssplitrow2[2:][0]
				healthchecklist.append(tempdict)
	except IndexError:
		print 'Could not connect to device ' + deviceip
		try:
			sshnet_connect.disconnect()
		except:
			pass
	except Exception as e:
		print 'Error while running health check with ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			pass
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting script'
		try:
			sshnet_connect.disconnect()
		except:
			pass
	try:
		sshnet_connect.disconnect()
		sys.exit()
	except:
		pass			
			
#########################################
print ''
print 'Switchport Manager'
print '############################################################'
print 'The purpose of this tool is to use a XLSX import to update'
print 'a ports vlan based on the vendor mac address and/or restart'
print 'the port based on the mac address.'
print 'Please fill in the config tab on the templated XLSX sheet.'
print '############################################################'
print ''
print '----Questions that need answering----'
excelfilelocation = raw_input('File to load the excel data from (e.g. C:/Python27/sw-config.xlsx):')
if excelfilelocation == '':
	excelfilelocation = 'C:/Python27/sw-config.xlsx'
excelfilelocation = excelfilelocation.replace('"', '')
# Load Configuration Variables
configdict = {}
for configvariables in xlhelper.sheet_to_dict(excelfilelocation,'Config'):
	try:
		configvar = configvariables.get('Variable').encode('utf-8')
		configval = configvariables.get('Value').encode('utf-8')
	except:
		configvar = configvariables.get('Variable')
		configval = configvariables.get('Value')
	configdict[configvar] = configval
# Username Variables/Questions
sshusername = configdict.get('Username')
if 'NA' == sshusername:
	sshusername = raw_input('What is the username you will use to login to the devices?:')
sshpassword = configdict.get('Password')
if 'NA' == sshpassword:
	sshpassword = getpass.getpass('What is the password you will use to login to the devices?:')
enablesecret = configdict.get('EnableSecret')
if 'NA' == enablesecret:
	enablesecret = getpass.getpass('What is the enable password you will use to access the devices?:')
# Rest of the Config Variables
loglocation = configdict.get('LogLocation')
if loglocation == None:
	loglocation = r'C:\Scripts\SWManager\Log'
vendormac = configdict.get('VendorMAC')
if vendormac == None:
	vendormac = raw_input('Please enter the first 6 characters of the vendor mac you want to match on?')
vendorvlan = configdict.get('VendorVLAN')
if vendorvlan == None:
	vendorvlan = raw_input('Please enter the vlan number you want to modify the ports to?')
vendorvlan = str(vendorvlan)
if ',' in vendorvlan:
	vendorvlan = vendorvlan.split(',')
vendorvlanmod = vendorvlan
vendortemplate = configdict.get('VendorTemplate')
if not vendortemplate == None:
	vendortemplate = str(vendortemplate)
exportlocation = configdict.get('ExportLocation')
if exportlocation == None:
	exportlocation = raw_input('Please enter the file path for the report (e.g. C:\Scripts\SWManager):')
	if exportlocation == '':
		exportlocation = 'C:\Scripts\SWManager'
exportlocation = str(exportlocation)
# Report Header Style
HeaderFont = Font(bold=True)
HeaderFont.size = 12
HeaderStyle = NamedStyle(name='BoldHeader')
HeaderStyle.font = HeaderFont
#
modifyvlan = 0
devicehostnames = []
finalinterfacelist = []
devicelist = []
healthchecklist = []
tempfilelist = []
for devices in xlhelper.sheet_to_dict(excelfilelocation,'Device IPs'):
	devicelist.append(devices)
# Create Log folder if its missing
newinstall = 0
if not os.path.exists(loglocation):
	os.makedirs(loglocation)
	newinstall = 1
# Install MAC DB into RAM
try:
	maclookupdburl = "http://standards-oui.ieee.org/oui.txt"
	maclookupfilename = 'oui.txt'
	if not os.path.isfile(maclookupfilename):
		DOWNLOAD_FILE(maclookupdburl, maclookupfilename)
	maclookupdbo = open(maclookupfilename)
	maclookupdb = maclookupdbo.readlines()
	maclookupdbo.close()
	skipmac = 0
except Exception as e:
	skipmac = 1
	print 'Could not load MAC database. Error is ' + str(e)		
#### Menu 
print ''
print '#################################################'
print '###                                           ###'
print '###        Please select an option below      ###'
print '###                                           ###'
print '###  1. Restart all vendor mac ports          ###'
print '###  2. Report VLAN for all vendor mac ports  ###'
print '###  3. Update VLAN for all vendor mac ports  ###'
print '###  4. Change the vendor mac address         ###'
print '###  5. Update the template on the port       ###'
print '###  6. Find devices in VLAN without an IP    ###'
print '###  7. Update devices in VLAN without an IP  ###'
print '###  8. Export Report to CSV of assignment    ###'
print '###  9. Interface Health Check                ###'
print '###  10. Exit                                 ###'
print '###                                           ###'
print '#################################################'
print ''
menuoption = raw_input('Selection (1-10)?:')
if menuoption == '':
	menuoption = 10
menuoption = int(menuoption)
# Modify VLAN option
if menuoption == 3:
	modifyvlan = 1
	if isinstance(vendorvlan, list):
		print '**************************************************'
		print '* Please select a VLAN in the list to change to: *'
		print '**************************************************'
		rowc = 1
		for v in vendorvlan:
			print str(rowc) + '. ' + v
			rowc = rowc + 1
		vendorvlanmod = raw_input('Please enter the vlan number you want to modify the ports to?: ')
else:
	modifyvlan = 0
if menuoption == 4:
		vendormac = raw_input('Please enter the first 6 characters of the vendor mac you want to match on?: ')
# Start



while (menuoption < 10):
	if menuoption == 1:
		threadstartime = datetime.now()
		if __name__ == "__main__":
			# Start Threads
			print 'Starting restart of all devices based on vendor mac'
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=RestartPort, args=(device,vendormac))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
		print 'Successfully restarted all ports based on vendor mac'
		print 'Successfully changed or viewed the vlan on all ports based on vendor mac'
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
	if menuoption == 2 or menuoption == 3:
		threadstartime = datetime.now()
		if __name__ == "__main__":	
		# Start Threads
			print 'Starting comparison/update of switchport based on vendor mac'
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=SetVLAN, args=(device,vendormac,vendorvlan,vendorvlanmod,vendortemplate))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
		print 'Successfully changed or viewed the vlan on all ports based on vendor mac'
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
	if menuoption == 5:
		threadstartime = datetime.now()
		if __name__ == "__main__":
		# Start Threads
			print 'Starting comparison/update of template to vlan'
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=SetTemplate, args=(device,vendorvlan,vendortemplate))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
		print 'Successfully changed or viewed the vlan on all ports based on vendor mac'
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
	if menuoption == 6:
		modifyipport = 0
		modifyvendorvlan = vendorvlan
		threadstartime = datetime.now()
		if __name__ == "__main__":
		# Start Threads
			print 'Looking for devices in the specified VLAN without an IP in the ARP table'
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=MacIPCompare, args=(device,vendorvlan,modifyipport,modifyvendorvlan))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
		print 'Successfully viewed all devices with mac addresses without IPs'
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
		modifyipport = 0
	if menuoption == 7:
		modifyipport = 1
		modifyvendorvlan = raw_input('What is the VLAN ID that you would like to move the mac addresses that are missing IPs into?: ')
		threadstartime = datetime.now()
		if __name__ == "__main__":
		# Start Threads
			print 'Updating devices in the specified VLAN without an IP in the ARP table'
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=MacIPCompare, args=(device,vendorvlan,modifyipport,modifyvendorvlan))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
		print 'Successfully changed the vlan on all ports without an IP'
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
		modifyipport = 0
	if menuoption == 8:
		if __name__ == "__main__":
		# Start Threads
			print 'Starting to gather data on switches'
			threadstartime = datetime.now()
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=ExportVLANs, args=(device,))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
			try:
				print 'Starting export to XLS of all switchports'
				wb = Workbook()
				today = date.today()
				datenum = today.strftime('%m%d%Y')
				dest_filename = 'VLAN-Report-' + datenum + '.xlsx'
				dest_path = exportlocation + '\\' + dest_filename
				# Multiple Devices Report (Separate Tabbed) WIP
				dcount = 0
				for d in devicehostnames:
					dcount = dcount + 1
				#
				ws1 = wb.active
				# Continue on with work
				ws1.title = "VLAN Export"
				ws1.append(['Hostname','Interface','VLAN','Port Status','MacAddress','MacVendor','IPAddress','VRF','Template','POE Status','POE Watts','POE Device','CDP Name','CDP Platform'])
				startrow = 2
				for row in finalinterfacelist:
					if row.get('MacCompany') == 'Unknown':
						intmac = row.get('MacAddress')
						try:
							mac_company_mac = str(intmac[0:7].replace('.','')).upper()
							for line in maclookupdb:
								if line.startswith(mac_company_mac):
									linev = line.replace('\n','').replace('\t',' ')
									maccompany = re.search(r'^[A-Z0-9]{6}\s+\(base 16\)\s+(.*)',linev).group(1)
								if maccompany == '' or maccompany == None:
									maccompany = 'Unknown'
						except:
							maccompany = 'Unknown'
					else:
						maccompany = row.get('MacCompany')
					# Add to workbook
					ws1['A' + str(startrow)] = row.get('Hostname')
					ws1['B' + str(startrow)] = row.get('Interface')
					ws1['C' + str(startrow)] = row.get('VLAN')
					ws1['D' + str(startrow)] = row.get('Status')
					ws1['E' + str(startrow)] = row.get('MacAddress')
					ws1['F' + str(startrow)] = maccompany
					ws1['G' + str(startrow)] = row.get('IPAddress')
					ws1['H' + str(startrow)] = row.get('VRF')
					ws1['I' + str(startrow)] = row.get('Template')
					ws1['J' + str(startrow)] = row.get('POEStatus')
					ws1['K' + str(startrow)] = row.get('POEWatts')
					ws1['L' + str(startrow)] = row.get('POEDevice')
					ws1['M' + str(startrow)] = row.get('CDPHostname')
					ws1['N' + str(startrow)] = row.get('CDPPlatform')
					startrow = startrow + 1
				wb.add_named_style(HeaderStyle)
				# Set Column Width
				for col in ws1.columns:
					 max_length = 0
					 column = col[0].column # Get the column name
					 for cell in col:
						 try: # Necessary to avoid error on empty cells
							 if len(str(cell.value)) > max_length:
								 max_length = len(cell.value)
						 except:
							 pass
					 adjusted_width = (max_length + 2) * 1.2
					 ws1.column_dimensions[column].width = adjusted_width
				# Set styles on header row
				for cell in ws1["1:1"]:
					cell.style = 'BoldHeader'
				wb.save(filename = dest_path)
				threadendtime = datetime.now()
				# Sort/filter XLSX, done through win32com
				sortq = raw_input('Would you like the XLSX file sorted/filtered? It requires Excel/Windows (Y/N): ')
				if 'y' in sortq.lower():
					excel = win32com.client.Dispatch("Excel.Application")
					wb = excel.Workbooks.Open(dest_path)
					ws = wb.Worksheets('VLAN Export')
					ws.Range('A2:N50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
					ws.Range('A1:N1').AutoFilter(1)
					wb.Save()
					excel.Application.Quit()
				print 'Successfully created VLAN Report'
				threadtime = threadendtime - threadstartime
				threadsec = threadtime.seconds
				print 'Elapsed time ' + str(threadsec) + ' seconds.'
			except Exception as e:
				print 'Error creating VLAN Report. Error is ' + str(e)
		print 'Successfully exported vlans on all ports to xlsx'
	if menuoption == 9:
		if __name__ == "__main__":
		# Start Threads
			print 'Starting to gather interface health data on switches'
			threadstartime = datetime.now()
			for device in devicelist:	
				deviceip = device.get('IP').encode('utf-8')
				t = threading.Thread(target=HealthCheck, args=(device,))
				t.start()
			main_thread = threading.currentThread()
			# Join All Threads
			for it_thread in threading.enumerate():
				if it_thread != main_thread:
					it_thread.join()
			print 'Exporting Health Reports'
			wb = Workbook()
			today = date.today()
			datenum = today.strftime('%m%d%Y')
			dest_filename = 'Health-Check-Report-' + datenum + '.xlsx'
			dest_path = exportlocation + '\\' + dest_filename
			ws1 = wb.active
			# Continue on with work
			ws1.title = "Health Check"
			ws1.append(['Hostname','Error','Description'])
			startrow = 2
			for row in healthchecklist:
				ws1['A' + str(startrow)] = row.get('Hostname')
				ws1['B' + str(startrow)] = row.get('Error')
				ws1['C' + str(startrow)] = row.get('Description')
				startrow = startrow + 1	
			wb.add_named_style(HeaderStyle)
			# Set styles on header row
			for cell in ws1["1:1"]:
				cell.style = 'BoldHeader'
			wb.save(filename = dest_path)
			sortq = raw_input('Would you like the XLSX file sorted/filtered? It requires Excel/Windows (Y/N): ')
			if 'y' in sortq.lower():
				excel = win32com.client.Dispatch("Excel.Application")
				wb = excel.Workbooks.Open(dest_path)
				ws = wb.Worksheets('Health Check')
				ws.Range('A2:C50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
				ws.Range('A1:C1').AutoFilter(1)
				wb.Save()
				excel.Application.Quit()
			print 'Successfully created Health Check Report'
	print ''
	print '#################################################'
	print '###                                           ###'
	print '###        Please select an option below      ###'
	print '###                                           ###'
	print '###  1. Restart all vendor mac ports          ###'
	print '###  2. Report VLAN for all vendor mac ports  ###'
	print '###  3. Update VLAN for all vendor mac ports  ###'
	print '###  4. Change the vendor mac address         ###'
	print '###  5. Update the template on the port       ###'
	print '###  6. Find devices in VLAN without an IP    ###'
	print '###  7. Update devices in VLAN without an IP  ###'
	print '###  8. Export Report to CSV of assignment    ###'
	print '###  9. Interface Health Check                ###'
	print '###  10. Exit                                 ###'
	print '###                                           ###'
	print '#################################################'
	print ''
	menuoption = raw_input('Selection (1-10)?:')
	if menuoption == '':
		menuoption = 10
	menuoption = int(menuoption)
	if menuoption == 3:
		modifyvlan = 1
		if isinstance(vendorvlan, list):
			print '**************************************************'
			print '* Please select a VLAN in the list to change to: *'
			print '**************************************************'
			rowc = 1
			for v in vendorvlan:
				print str(rowc) + '. ' + v
				rowc = rowc + 1
			vendorvlanmod = raw_input('Please enter the vlan number you want to modify the ports to?: ')
	else:
		modifyvlan = 0
	if menuoption == 4:
		vendormac = raw_input('Please enter the first 6 characters of the vendor mac you want to match on?: ')
print ''
print 'Exiting...'
print 'Thanks for playing'