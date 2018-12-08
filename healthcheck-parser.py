#!/usr/bin/env python

'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL netmiko (pip install netmiko)
INSTALL textfsm (pip install textfsm)
INSTALL openpyxl (pip install openpyxl)
INSTALL fileinput (pip install fileinput)
INSTALL xlhelper (python -m pip install git+git://github.com/routeallthings/xlhelper.git)
'''

#Module Imports (Native)
import re
import os
import unicodedata
import csv
import threading
import time
import sys
from datetime import datetime
from datetime import date

#Module Imports (Non-Native)
# Itertools
from string import ascii_lowercase
import itertools

#Requests
try:
	import requests
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('Requests module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in requestsinstallstatus.upper() or "YES" in requestsinstallstatus.upper():
		os.system('python -m pip install requests')
		import requests
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of requests. Please install manually and retry"
		sys.exit()
#OpenPYXL
try:
	from openpyxl import load_workbook
	from openpyxl import workbook
	from openpyxl import Workbook
	from openpyxl.styles import Font, NamedStyle
except ImportError:
	openpyxlinstallstatus = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in openpyxlinstallstatus.lower():
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
		from openpyxl import workbook
		from openpyxl import Workbook
		from openpyxl.styles import Font, NamedStyle
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of openpyxl. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		
		
#######################################
#Functions
def ITER_ALL_STRINGS():
    for size in itertools.count(1):
        for s in itertools.product(ascii_lowercase, repeat=size):
            yield "".join(s)

def LABEL_GEN():
    for s in gen:
        return s			

def DOWNLOAD_FILE(url,saveas):
    # NOTE the stream=True parameter
    r = requests.get(url, stream=True)
    with open(saveas, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024): 
            if chunk: # filter out keep-alive new chunks
                f.write(chunk)	
				
#########################################
print ''
print 'ISE Export Information'
print '############################################################'
print 'The purpose of this tool is to use a csv import to parse'
print 'data from ISE in order to create a better report for use.'
print 'Using this information you can load an additional data'
print 'into ISE to create your profiling rules.'
print '############################################################'
print ''
print '----Questions that need answering----'
exportlocation = raw_input('Please enter the file path for the report (e.g. C:\Scripts\SWManager):')
if exportlocation == '':
	exportlocation = 'C:\Scripts\SWManager'
exportlocation = str(exportlocation)
importlocation = raw_input('Please enter the import folder path (it will check all subfolders (e.g. C:\Scripts\SWManager\HC)):')
if importlocation == '':
	importlocation = 'C:\Scripts\SWManager\HC'
importlocation = str(importlocation)
# Report Header Style
HeaderFont = Font(bold=True)
HeaderFont.size = 12
HeaderStyle = NamedStyle(name='BoldHeader')
HeaderStyle.font = HeaderFont
# Create Exportlocation folder if its missing
if not os.path.exists(exportlocation):
	os.makedirs(exportlocation)
#### Menu 
print ''
print '#################################################'
print '###                                           ###'
print '###        Please select an option below      ###'
print '###                                           ###'
print '###  1. Export health check TXT to XLSX       ###'
print '###  2. PLACEHOLDER                           ###'
print '###  3. PLACEHOLDER                           ###'
print '###  4. Exit                                  ###'
print '###                                           ###'
print '#################################################'
print ''
menuoption = raw_input('Selection (1-4)?:')
menuoption = int(menuoption)
if menuoption == '':
	menuoption = 4
# Start
poelist = []
poehostnamelist = []
droplist = []
drophostnamelist = []
while (menuoption < 4):
	if menuoption == 1:
		# Get Folder Paths for first loop through directories
		rootfolderlist = os.listdir(importlocation)
		for folder in rootfolderlist:
			subfolder = importlocation + '\\' + folder
			subfolderlist = os.listdir(subfolder)
			for file in subfolderlist:
				if '.txt' in file:
					filepath = subfolder + '\\' + file
					fopen = open(filepath, "r")
					if 'POE' in file:
						filere = re.search('(\S+)-POE-(.*).txt',file)
						timestamp = datetime.strptime(filere.group(2),'%m%d%Y-%H%M%S')
						hostname = filere.group(1)
						hostnameduplicate = 0
						# Create POE hostname list
						for hostnamelist in poehostnamelist:
							if hostnamelist == hostname:
								hostnameduplicate = 1
								break
						if hostnameduplicate == 0:
							poehostnamelist.append(hostname)
						for line in fopen:
							if '1800' in line:
								poedict = {}
								fullpoere = re.search('(\S+)\s+(\S+)\s+(\S+)\s+(\S+).*',line)
								switchnum = fullpoere.group(1)
								switchused = fullpoere.group(3)
								switchremain = fullpoere.group(4)
								poedict['hostname'] = hostname
								poedict['timestamp'] = timestamp
								poedict['number'] = switchnum
								poedict['powerused'] = switchused
								poedict['powerremaining'] = switchremain
								poelist.append(poedict)
					if 'ShowDrops' in file:
						filere = re.search('(\S+)-ShowDrops-(.*).txt',file)
						timestamp = datetime.strptime(filere.group(2),'%m%d%Y-%H%M%S')
						hostname = filere.group(1)
						filedroplist = []
						filedroplistinterfaces = []
						count = 1
						# Create DROP hostname list
						hostnameduplicate = 0
						for hostnamelist in drophostnamelist:
							if hostnamelist == hostname:
								hostnameduplicate = 1
								break
						if hostnameduplicate == 0:
							drophostnamelist.append(hostname)
						# Gather
						for line in fopen:
							linedict = {}
							if 'line protocol' in line:
								intnamere = re.search('(\S+).*',line)
								intname = intnamere.group(1)
								linedict['number'] = count
								linedict['interface'] = intname
								filedroplist.append(linedict)
								filedroplistinterfaces.append(intname)
							if 'output drops' in line:
								outputdropre = re.search('.*Total output drops: (\d+).*',line)
								outputdrop = outputdropre.group(1)
								linedict['number'] = count
								linedict['drops'] = outputdrop
								filedroplist.append(linedict)
								count = count + 1
						for interfaces in filedroplistinterfaces:
							newintdict = {}
							for intdict in filedroplist:
								if intdict.get('interface') == interfaces:
									intkey = intdict.get('number')
									break
							for intdict in filedroplist:
								if intdict.get('number') == intkey and not intdict.get('drops') == '':
									intdrops = intdict.get('drops')
							newintdict['hostname'] = hostname
							newintdict['timestamp'] = timestamp
							newintdict['interface'] = interfaces
							newintdict['drops'] = intdrops
							droplist.append(newintdict)
		# Start work
		threadstartime = datetime.now()
		print 'Starting Export of health check data to XLSX'
		# Create XLSX for POE
		wb = Workbook()
		today = date.today()
		datenum = today.strftime('%m%d%Y')
		dest_filename = 'HealthCheck-POE-Report-' + datenum + '.xlsx'
		dest_path = exportlocation + '\\' + dest_filename
		ws1 = wb.active
		# Continue on with work
		ws1.title = "HC Export POE"
		ws1.append(['Hostname','Timestamp','Switch Number','Power Used (Watts)','Power Remaining (Watts)'])
		startrow = 2
		for poehost in poehostnamelist:
			for p in poelist:
				if poehost == p.get('hostname'):
					ws1['A' + str(startrow)] = p.get('hostname')
					ws1['B' + str(startrow)] = p.get('timestamp')
					ws1['C' + str(startrow)] = int(p.get('number'))
					ws1['D' + str(startrow)] = float(p.get('powerused'))
					ws1['E' + str(startrow)] = float(p.get('powerremaining'))
					startrow = startrow + 1
		wb.add_named_style(HeaderStyle)
		# Set Column Width
		for col in ws1.columns:
			 max_length = 0
			 column = col[0].column # Get the column name
			 for cell in col:
				 try: # Necessary to avoid error on empty cells
					 if len(str(cell.value)) > max_length:
						 max_length = len(cell.value)
				 except:
					 pass
			 adjusted_width = (max_length + 2) * 1.2
			 ws1.column_dimensions[column].width = adjusted_width
		# Set styles on header row
		for cell in ws1["1:1"]:
			cell.style = 'BoldHeader'
		wb.save(filename = dest_path)
		# Create XLSX for Drops
		wb2 = Workbook()
		today = date.today()
		datenum = today.strftime('%m%d%Y')
		dest_filename = 'HealthCheck-Drops-Report-' + datenum + '.xlsx'
		dest_path = exportlocation + '\\' + dest_filename
		ws1 = wb2.active
		# Continue on with work
		ws1.title = "HC Export Drops"
		ws1.append(['Hostname','Timestamp','Interface','Drops (in Bytes)'])
		startrow = 2
		for drophost in drophostnamelist:
			for d in droplist:
				if drophost == d.get('hostname'):
					ws1['A' + str(startrow)] = d.get('hostname')
					ws1['B' + str(startrow)] = d.get('timestamp')
					ws1['C' + str(startrow)] = d.get('interface')
					ws1['D' + str(startrow)] = int(d.get('drops'))
					startrow = startrow + 1
		wb2.add_named_style(HeaderStyle)
		# Set Column Width
		for col in ws1.columns:
			 max_length = 0
			 column = col[0].column # Get the column name
			 for cell in col:
				 try: # Necessary to avoid error on empty cells
					 if len(str(cell.value)) > max_length:
						 max_length = len(cell.value)
				 except:
					 pass
			 adjusted_width = (max_length + 2) * 1.2
			 ws1.column_dimensions[column].width = adjusted_width
		# Set styles on header row
		for cell in ws1["1:1"]:
			cell.style = 'BoldHeader'
		wb2.save(filename = dest_path)
		threadendtime = datetime.now()
		threadtime = threadendtime - threadstartime
		threadsec = threadtime.seconds
		print 'Elapsed time ' + str(threadsec) + ' seconds.'
							
	print ''
	print '#################################################'
	print '###                                           ###'
	print '###        Please select an option below      ###'
	print '###                                           ###'
	print '###  1. Export health check TXT to XLSX       ###'
	print '###  2. PLACEHOLDER                           ###'
	print '###  3. PLACEHOLDER                           ###'
	print '###  4. Exit                                  ###'
	print '###                                           ###'
	print '#################################################'
	print ''
	menuoption = raw_input('Selection (1-4)?:')
	menuoption = int(menuoption)
	if menuoption == '':
		menuoption = 4
print ''
print 'Exiting...'
print 'Thanks for playing'
